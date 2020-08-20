import dropbox
import openpyxl as pyxl
from progress.bar import IncrementalBar


def fillWithSharedLinks(accessToken, imgFolders, xlPath, savePath):
    # create the client object
    dbx = dropbox.Dropbox(accessToken)

    img_link_dict = dict()

    print("Getting Links .....")

    # loop through the folders
    for im in range(0, len(imgFolders)):

        imgFolder = imgFolders[im]

        # create a list of all images in the folder
        imgs = getAllFiles(dbx, imgFolder)

        # iterate through the list of images and create shared links (store the codes and links in a dict)
        bar = IncrementalBar(('Folder %d/%d' % (im + 1, len(imgFolders))), max=len(imgs))

        for entry in imgs:
            nm = entry.name
            pth = imgFolder + "/" + nm

            # check if a link already exists and retrieve it, if not create one
            lnkdata = dbx.sharing_list_shared_links(pth, direct_only=True)

            if len(lnkdata.links) == 0:
                metadata = dbx.sharing_create_shared_link_with_settings(pth)
                lnk = metadata.url
            else:
                lnk = lnkdata.links[0].url

            img_link_dict[nm[:-4]] = lnk

            bar.next()
        bar.finish()

    # open the excel file and get the sheet
    workBook = pyxl.load_workbook(xlPath)
    sheet = workBook['List']

    # iterate through
    print("Filling Spreadsheet .....")
    for row in range(2, sheet.max_row):
        # get the code
        code = str(sheet.cell(row, 1).value)

        if not img_link_dict.keys().__contains__(code):
            continue

        # get the link
        link = img_link_dict[code]

        # insert the link in the sheet
        sheet.cell(row, 3).value = link

    # save the excel sheet
    savenm = xlPath.split("\\")[-1].split(".")[0] + ' - With Links.xlsx'
    workBook.save(savePath + "\\" + savenm)


def getAllFiles(dbx, imgFolder):

    # initialize
    allFiles = []

    # get files
    currFiles = dbx.files_list_folder(imgFolder)

    # loop through the entries in the current list
    def process_entries(entries):
        for entry in entries:
            if isinstance(entry, dropbox.files.FileMetadata):
                allFiles.append(entry)

    # process current entries
    process_entries(currFiles.entries)

    # loop until you get all files
    while currFiles.has_more:
        currFiles = dbx.files_list_folder_continue(currFiles.cursor)

        # process again
        process_entries(currFiles.entries)

    # return
    return allFiles


def main():
    accessToken = "INSERT ACCESSTOKEN HERE"

    imgFolders = ["INSERT IMAGEFOLDERS NAME HERE"]

    xlPath = r"INSERT EXCEL PATH HERE"

    savePath = r"INSERT SAVING PATH HERE"
    return fillWithSharedLinks(accessToken, imgFolders, xlPath, savePath)


if __name__ == "__main__":
    main()
